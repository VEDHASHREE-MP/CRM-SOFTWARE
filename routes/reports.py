from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from database import get_db
from datetime import datetime
import psycopg2.extras
import traceback
import io

reports_bp = Blueprint('reports', __name__)


def get_cursor(db):
    return db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)


def get_user(user_id):
    db  = get_db()
    cur = get_cursor(db)
    cur.execute('SELECT id, role FROM users WHERE id = %s', (user_id,))
    user = cur.fetchone()
    db.close()
    return user


def date_params():
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')
    return date_from, date_to


def build_date_filter(col, date_from, date_to, prefix='AND'):
    clauses = []
    params  = []
    if date_from:
        clauses.append(f'{prefix} SUBSTR({col}::text, 1, 10) >= %s')
        params.append(date_from)
        prefix = 'AND'
    if date_to:
        clauses.append(f'{prefix} SUBSTR({col}::text, 1, 10) <= %s')
        params.append(date_to)
    return ' '.join(clauses), params


# ── Summary (cards) ───────────────────────────────────────────────────────────

@reports_bp.route('/summary', methods=['GET'])
@jwt_required()
def summary():
    user_id  = get_jwt_identity()
    user     = get_user(user_id)
    is_admin = user['role'] == 'admin'
    date_from, date_to = date_params()

    db  = get_db()
    cur = get_cursor(db)

    enq_date_frag,  enq_date_params  = build_date_filter('e.created_at', date_from, date_to)
    enq_staff_frag  = '' if is_admin else 'AND e.assigned_to = %s'
    enq_base_params = [] if is_admin else [user_id]

    def enq_count(extra_sql='', extra_params=[]):
        cur.execute(
            f'SELECT COUNT(*) AS c FROM enquiries e WHERE 1=1 {enq_staff_frag} {enq_date_frag} {extra_sql}',
            enq_base_params + enq_date_params + extra_params
        )
        return cur.fetchone()['c']

    total_enq  = enq_count()
    new_enq    = enq_count('AND e.status = %s', ['New'])
    converted  = enq_count('AND e.status = %s', ['Converted'])
    followup   = enq_count('AND e.status = %s', ['Follow-up'])
    closed     = enq_count('AND e.status = %s', ['Closed'])

    sess_date_frag,  sess_date_params  = build_date_filter('s.created_at', date_from, date_to)
    sess_staff_frag  = '' if is_admin else 'AND s.assigned_staff = %s'
    sess_base_params = [] if is_admin else [user_id]

    def sess_count(extra_sql='', extra_params=[]):
        cur.execute(
            f'SELECT COUNT(*) AS c FROM sessions s WHERE 1=1 {sess_staff_frag} {sess_date_frag} {extra_sql}',
            sess_base_params + sess_date_params + extra_params
        )
        return cur.fetchone()['c']

    total_sess  = sess_count()
    completed_s = sess_count('AND s.status = %s', ['Completed'])
    scheduled_s = sess_count('AND s.status = %s', ['Scheduled'])
    cancelled_s = sess_count('AND s.status = %s', ['Cancelled'])

    billing = {}
    if is_admin:
        inv_date_frag, inv_date_params = build_date_filter('i.created_at',   date_from, date_to)
        pay_date_frag, pay_date_params = build_date_filter('p.payment_date', date_from, date_to)

        cur.execute(
            f"SELECT COALESCE(SUM(i.total_amount),0) AS s FROM invoices i WHERE i.status != 'Cancelled' {inv_date_frag}",
            inv_date_params
        )
        total_invoiced = float(cur.fetchone()['s'])

        cur.execute(
            f"SELECT COALESCE(SUM(p.amount),0) AS s FROM payments p WHERE 1=1 {pay_date_frag}",
            pay_date_params
        )
        total_paid = float(cur.fetchone()['s'])

        cur.execute(
            f"SELECT COALESCE(SUM(i.due_amount),0) AS s FROM invoices i WHERE i.status NOT IN ('Paid','Cancelled') {inv_date_frag}",
            inv_date_params
        )
        total_due = float(cur.fetchone()['s'])

        cur.execute(
            f"SELECT COUNT(*) AS c FROM invoices i WHERE 1=1 {inv_date_frag}",
            inv_date_params
        )
        total_inv_count = cur.fetchone()['c']

        billing = {
            'total_invoiced': round(total_invoiced, 2),
            'total_paid':     round(total_paid, 2),
            'total_due':      round(total_due, 2),
            'invoice_count':  total_inv_count,
        }

    customers = {}
    if is_admin:
        cur.execute('''
            SELECT loyalty_tier, SUM(cnt) AS cnt FROM (
                SELECT loyalty_tier, COUNT(*) AS cnt
                FROM enquiries WHERE status = 'Converted'
                GROUP BY loyalty_tier
                UNION ALL
                SELECT loyalty_tier, COUNT(*) AS cnt
                FROM invoice_customers
                GROUP BY loyalty_tier
            ) sub GROUP BY loyalty_tier
        ''')
        tier_rows  = cur.fetchall()
        tier_map   = {r['loyalty_tier']: r['cnt'] for r in tier_rows}
        total_cust = sum(tier_map.values())
        customers  = {
            'total':   total_cust,
            'vip':     tier_map.get('VIP',     0),
            'loyal':   tier_map.get('Loyal',   0),
            'regular': tier_map.get('Regular', 0),
            'new':     tier_map.get('New',     0),
        }

    db.close()
    return jsonify({
        'enquiries': {
            'total': total_enq, 'new': new_enq, 'converted': converted,
            'follow_up': followup, 'closed': closed,
            'conversion_rate': round(converted / total_enq * 100, 1) if total_enq else 0
        },
        'sessions': {
            'total': total_sess, 'completed': completed_s,
            'scheduled': scheduled_s, 'cancelled': cancelled_s,
            'completion_rate': round(completed_s / total_sess * 100, 1) if total_sess else 0
        },
        'billing':   billing,
        'customers': customers,
    }), 200


# ── Chart data ────────────────────────────────────────────────────────────────

@reports_bp.route('/charts', methods=['GET'])
@jwt_required()
def charts():
    user_id  = get_jwt_identity()
    user     = get_user(user_id)
    is_admin = user['role'] == 'admin'
    date_from, date_to = date_params()

    db  = get_db()
    cur = get_cursor(db)

    enq_date_frag,  enq_date_params  = build_date_filter('e.created_at', date_from, date_to)
    sess_date_frag, sess_date_params = build_date_filter('s.created_at', date_from, date_to)

    enq_staff_frag   = '' if is_admin else 'AND e.assigned_to = %s'
    enq_base_params  = [] if is_admin else [user_id]
    sess_staff_frag  = '' if is_admin else 'AND s.assigned_staff = %s'
    sess_base_params = [] if is_admin else [user_id]

    cur.execute(
        f'SELECT e.status, COUNT(*) AS cnt FROM enquiries e WHERE 1=1 {enq_staff_frag} {enq_date_frag} GROUP BY e.status',
        enq_base_params + enq_date_params
    )
    enq_by_status = [dict(r) for r in cur.fetchall()]

    cur.execute(
        f'''SELECT e.source, COUNT(*) AS cnt FROM enquiries e
            WHERE 1=1 {enq_staff_frag} {enq_date_frag}
            GROUP BY e.source ORDER BY cnt DESC LIMIT 8''',
        enq_base_params + enq_date_params
    )
    enq_by_source = [dict(r) for r in cur.fetchall()]

    cur.execute(
        f'SELECT s.status, COUNT(*) AS cnt FROM sessions s WHERE 1=1 {sess_staff_frag} {sess_date_frag} GROUP BY s.status',
        sess_base_params + sess_date_params
    )
    sess_by_status = [dict(r) for r in cur.fetchall()]

    sess_by_staff = []
    if is_admin:
        cur.execute(
            f'''SELECT COALESCE(u.name, 'Unassigned') AS staff, COUNT(*) AS cnt
                FROM sessions s LEFT JOIN users u ON s.assigned_staff = u.id
                WHERE 1=1 {sess_date_frag}
                GROUP BY s.assigned_staff, u.name ORDER BY cnt DESC LIMIT 10''',
            sess_date_params
        )
        sess_by_staff = [dict(r) for r in cur.fetchall()]

    revenue_trend = []
    if is_admin:
        cur.execute('''
            SELECT to_char(payment_date::date, 'YYYY-MM') AS month,
                   ROUND(SUM(amount)::numeric, 2) AS revenue
            FROM payments
            GROUP BY month ORDER BY month DESC LIMIT 6
        ''')
        revenue_trend = list(reversed([dict(r) for r in cur.fetchall()]))

    cur.execute(
        f'''SELECT to_char(e.created_at::date, 'YYYY-MM') AS month, COUNT(*) AS cnt
            FROM enquiries e WHERE 1=1 {enq_staff_frag} {enq_date_frag}
            GROUP BY month ORDER BY month DESC LIMIT 6''',
        enq_base_params + enq_date_params
    )
    enq_trend = list(reversed([dict(r) for r in cur.fetchall()]))

    loyalty_breakdown = []
    if is_admin:
        cur.execute('''
            SELECT loyalty_tier AS tier, SUM(cnt) AS cnt FROM (
                SELECT loyalty_tier, COUNT(*) AS cnt
                FROM enquiries WHERE status = 'Converted'
                GROUP BY loyalty_tier
                UNION ALL
                SELECT loyalty_tier, COUNT(*) AS cnt
                FROM invoice_customers
                GROUP BY loyalty_tier
            ) sub GROUP BY loyalty_tier
            ORDER BY CASE loyalty_tier
                WHEN 'VIP'     THEN 1
                WHEN 'Loyal'   THEN 2
                WHEN 'Regular' THEN 3
                WHEN 'New'     THEN 4
                ELSE 5
            END
        ''')
        loyalty_breakdown = [{'tier': r['tier'], 'cnt': int(r['cnt'])} for r in cur.fetchall()]

    db.close()
    return jsonify({
        'enquiries_by_status': enq_by_status,
        'enquiries_by_source': enq_by_source,
        'sessions_by_status':  sess_by_status,
        'sessions_by_staff':   sess_by_staff,
        'revenue_trend':       revenue_trend,
        'enquiries_trend':     enq_trend,
        'loyalty_breakdown':   loyalty_breakdown,
    }), 200


# ── Excel Download ────────────────────────────────────────────────────────────

@reports_bp.route('/download', methods=['GET'])
@jwt_required()
def download_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        user_id  = get_jwt_identity()
        user     = get_user(user_id)
        is_admin = user['role'] == 'admin'
        date_from, date_to = date_params()

        db  = get_db()
        cur = get_cursor(db)

        wb = Workbook()

        HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
        HEADER_FILL = PatternFill('solid', fgColor='2563EB')
        TITLE_FONT  = Font(bold=True, size=13, color='1E293B')
        CENTER      = Alignment(horizontal='center', vertical='center')
        WRAP        = Alignment(wrap_text=True, vertical='top')
        THIN        = Side(style='thin', color='E2E8F0')
        BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ALT_FILL    = PatternFill('solid', fgColor='F8FAFC')

        def style_sheet(ws, headers, rows, sheet_title):
            ws.append([sheet_title])
            ws.cell(1, 1).font = TITLE_FONT
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(1, 1).alignment = CENTER
            if date_from or date_to:
                period = f"Period: {date_from or 'Start'} to {date_to or 'Today'}"
                ws.append([period])
                ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
                ws.cell(2, 1).alignment = CENTER
                ws.cell(2, 1).font = Font(italic=True, color='64748B', size=10)
                header_row = 3
            else:
                ws.append([])
                header_row = 3

            ws.append(headers)
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(header_row, col_idx)
                cell.font      = HEADER_FONT
                cell.fill      = HEADER_FILL
                cell.alignment = CENTER
                cell.border    = BORDER

            for row_idx, row in enumerate(rows, header_row + 1):
                ws.append(row)
                fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row_idx, col_idx)
                    cell.border    = BORDER
                    cell.alignment = WRAP
                    if fill.fill_type:
                        cell.fill = fill

            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    (len(str(ws.cell(r, col_idx).value or '')) for r in range(1, ws.max_row + 1)),
                    default=10
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
            ws.row_dimensions[header_row].height = 20
            return header_row

        # ── Sheet 1: Enquiries ──
        ws1 = wb.active
        ws1.title = 'Enquiries'
        enq_date_frag,  enq_date_params  = build_date_filter('e.created_at', date_from, date_to)
        enq_staff_frag  = '' if is_admin else 'AND e.assigned_to = %s'
        enq_base_params = [] if is_admin else [user_id]

        cur.execute(
            f'''SELECT e.name, e.phone, e.service, e.source, e.status,
                       e.notes, e.follow_up_date, u.name AS assigned_to, e.created_at
                FROM enquiries e LEFT JOIN users u ON e.assigned_to = u.id
                WHERE 1=1 {enq_staff_frag} {enq_date_frag} ORDER BY e.created_at DESC''',
            enq_base_params + enq_date_params
        )
        enq_rows = cur.fetchall()
        style_sheet(ws1,
            ['Name', 'Phone', 'Service', 'Source', 'Status', 'Notes',
             'Follow-up Date', 'Assigned To', 'Created At'],
            [[r['name'], r['phone'], r['service'], r['source'], r['status'],
              r['notes'] or '', r['follow_up_date'] or '',
              r['assigned_to'] or 'Unassigned', str(r['created_at'])[:10]] for r in enq_rows],
            'Enquiries Report'
        )

        # ── Sheet 2: Sessions ──
        ws2 = wb.create_sheet('Sessions')
        sess_date_frag,  sess_date_params  = build_date_filter('s.created_at', date_from, date_to)
        sess_staff_frag  = '' if is_admin else 'AND s.assigned_staff = %s'
        sess_base_params = [] if is_admin else [user_id]

        cur.execute(
            f'''SELECT s.customer_name, s.phone, s.service, s.session_date,
                       s.session_time, s.duration, s.mode,
                       s.location_or_link, s.status, u.name AS staff_name, s.notes
                FROM sessions s LEFT JOIN users u ON s.assigned_staff = u.id
                WHERE 1=1 {sess_staff_frag} {sess_date_frag} ORDER BY s.session_date DESC''',
            sess_base_params + sess_date_params
        )
        sess_rows = cur.fetchall()
        style_sheet(ws2,
            ['Customer', 'Phone', 'Service', 'Date', 'Time', 'Duration',
             'Mode', 'Location/Link', 'Status', 'Staff', 'Notes'],
            [[r['customer_name'], r['phone'], r['service'], r['session_date'],
              r['session_time'], r['duration'], r['mode'],
              r['location_or_link'] or '', r['status'],
              r['staff_name'] or 'Unassigned', r['notes'] or ''] for r in sess_rows],
            'Sessions Report'
        )

        # ── Sheet 3: Billing (admin only) ──
        if is_admin:
            ws3 = wb.create_sheet('Billing')
            inv_date_frag, inv_date_params = build_date_filter('i.created_at', date_from, date_to)
            cur.execute(
                f'''SELECT i.invoice_number, i.customer_name, i.phone,
                           i.total_amount, i.paid_amount, i.due_amount,
                           i.status, i.gst_rate, i.discount,
                           u.name AS created_by, i.created_at
                    FROM invoices i LEFT JOIN users u ON i.created_by = u.id
                    WHERE 1=1 {inv_date_frag} ORDER BY i.created_at DESC''',
                inv_date_params
            )
            inv_rows = cur.fetchall()
            style_sheet(ws3,
                ['Invoice #', 'Customer', 'Phone', 'Total (₹)', 'Paid (₹)', 'Due (₹)',
                 'Status', 'GST %', 'Discount (₹)', 'Created By', 'Date'],
                [[r['invoice_number'], r['customer_name'], r['phone'],
                  round(float(r['total_amount']), 2), round(float(r['paid_amount']), 2),
                  round(float(r['due_amount']), 2), r['status'],
                  float(r['gst_rate']), float(r['discount']),
                  r['created_by'] or '', str(r['created_at'])[:10]] for r in inv_rows],
                'Billing Report'
            )

        # ── Sheet 4: Customers (admin only) ──
        if is_admin:
            ws4 = wb.create_sheet('Customers')
            cur.execute('''
                SELECT e.name, e.phone, e.service, e.source, e.loyalty_tier,
                       COUNT(DISTINCT s.id) AS total_sessions,
                       SUM(CASE WHEN s.status = 'Completed' THEN 1 ELSE 0 END) AS completed,
                       COALESCE(SUM(DISTINCT CASE WHEN i.id IS NOT NULL THEN i.total_amount ELSE 0 END), 0) AS total_invoiced,
                       COALESCE(SUM(DISTINCT CASE WHEN i.id IS NOT NULL THEN i.paid_amount  ELSE 0 END), 0) AS total_paid,
                       COALESCE(SUM(DISTINCT CASE WHEN i.id IS NOT NULL THEN i.due_amount   ELSE 0 END), 0) AS total_due,
                       e.updated_at AS cust_since
                FROM enquiries e
                LEFT JOIN sessions s ON s.enquiry_id = e.id
                LEFT JOIN invoices i ON i.enquiry_id = e.id AND i.status != 'Cancelled'
                WHERE e.status = 'Converted'
                GROUP BY e.id ORDER BY e.name
            ''')
            enq_cust_rows = cur.fetchall()

            cur.execute('''
                SELECT ic.name, ic.phone, ic.loyalty_tier,
                       STRING_AGG(DISTINCT ii.description, ', ') AS descriptions,
                       COUNT(DISTINCT i.id)             AS total_invoices,
                       COALESCE(SUM(i.total_amount), 0) AS total_invoiced,
                       COALESCE(SUM(i.paid_amount),  0) AS total_paid,
                       COALESCE(SUM(i.due_amount),   0) AS total_due,
                       ic.created_at
                FROM invoice_customers ic
                LEFT JOIN invoices i
                    ON i.phone = ic.phone AND i.enquiry_id IS NULL AND i.status != 'Cancelled'
                LEFT JOIN invoice_items ii ON ii.invoice_id = i.id
                GROUP BY ic.id ORDER BY ic.name
            ''')
            inv_cust_rows = cur.fetchall()

            combined_rows = []
            for r in enq_cust_rows:
                combined_rows.append([
                    r['name'], r['phone'], 'Session Customer', r['service'], r['source'],
                    r['loyalty_tier'], int(r['total_sessions'] or 0), int(r['completed'] or 0),
                    round(float(r['total_invoiced']), 2), round(float(r['total_paid']), 2),
                    round(float(r['total_due']), 2), str(r['cust_since'])[:10],
                ])
            for r in inv_cust_rows:
                raw_desc   = r['descriptions'] or ''
                desc_parts = list(dict.fromkeys(d.strip() for d in raw_desc.split(',') if d.strip()))
                description = ', '.join(desc_parts[:3])
                if len(desc_parts) > 3:
                    description += f' (+{len(desc_parts) - 3} more)'
                combined_rows.append([
                    r['name'], r['phone'], 'Invoice Customer', description or '—', 'Invoice',
                    r['loyalty_tier'], 0, 0,
                    round(float(r['total_invoiced']), 2), round(float(r['total_paid']), 2),
                    round(float(r['total_due']), 2), str(r['created_at'])[:10],
                ])

            header_row = style_sheet(ws4,
                ['Name', 'Phone', 'Type', 'Service / Description', 'Source', 'Loyalty Tier',
                 'Total Sessions', 'Completed Sessions',
                 'Total Invoiced (₹)', 'Total Paid (₹)', 'Total Due (₹)', 'Customer Since'],
                combined_rows,
                'Customers Report'
            )
            SESSION_FILL = PatternFill('solid', fgColor='DBEAFE')
            INVOICE_FILL = PatternFill('solid', fgColor='FEF9C3')
            data_start   = 5 if (date_from or date_to) else 4
            for offset, row_data in enumerate(combined_rows):
                row_idx = data_start + offset
                fill    = SESSION_FILL if row_data[2] == 'Session Customer' else INVOICE_FILL
                for col_idx in range(1, 13):
                    ws4.cell(row_idx, col_idx).fill = fill

        db.close()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = datetime.now().strftime('%Y%m%d')
        filename = f'VirtualTech_Report_{date_str}.xlsx'

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500